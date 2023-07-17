#!/usr/bin/env python
#encoding:utf-8

# Copyright@SecuritySoftware Team(Chu Chen etc. github.com/secsoftware)
# If your project uses or refer to our code, please cite as follows. Thank you!
"""
Chu Chen, Pinghong Ren, Zhenhua Duan, Cong Tian, Xu Lu, and Bin Yu. 2023. 
SBDT: Search-Based Differential Testing of Certificate Parsers in SSL/TLS Implementations. 
In Proceedings of the 32nd ACM SIGSOFT International Symposium on Software Testing and Analysis (ISSTA ’23), July
17–21, 2023, Seattle, WA, USA. ACM, New York, NY, USA, 13 pages. 
https://doi.org/10.1145/3597926.3598110
""" 

import os
import time
import re
import json
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter


# -+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+-+
tlsimpl = ["OpenSSL", "ZCertificate", "GnuTLS"] 
dir_cp_tmp = "../cp_tmp/"  
dir_cp_uniform = "../cp_uniform/" 
cannot_parsed_by_tlsimpl = "cannot_parsed_by_" 
dir_result = "../result/" 
file_result = dir_result + "results-dtcp.xlsx" 
global_cert_list = [] 


def write_xlsx(path_to_xlsx, worksheet_name, data_dict, initial=False):
    """ To save results to a .xlsx file.
    @param path_to_xlsx: A .xlsx file and a path to it;
    @param worksheet_name: A worksheet name;
    @param data_dict: Data to store and its type is dict;
    @param initial: A boolean indicating whether to initialize the worksheet;
    @return: A worksheet containing new data. 
    """
    if initial:
        pddf = pd.DataFrame(data_dict) 
        pddf.to_excel(path_to_xlsx, 
                      sheet_name=worksheet_name,
                      index=False,
                      header=True)
        workbook = openpyxl.load_workbook(path_to_xlsx)
        worksheet = workbook[worksheet_name]
        column_width = (pddf.columns.to_series().apply(lambda x: len(x)).values)
        for i, width in enumerate(column_width, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width + 2
        workbook.save(path_to_xlsx)
        return()
    writer = pd.ExcelWriter(path_to_xlsx, engine='openpyxl')
    book = openpyxl.load_workbook(writer.path)
    writer.book = book
    pddf = pd.DataFrame(data_dict)
    pddf.to_excel(excel_writer=writer, 
                  sheet_name=worksheet_name, 
                  index=False,
                  header=True)
    writer.save()
    writer.close()
    
    workbook = openpyxl.load_workbook(path_to_xlsx)
    worksheet = workbook[worksheet_name]
    column_width = (pddf.columns.to_series().apply(lambda x: len(x)).values)
    for i, width in enumerate(column_width, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = width + 2
    workbook.save(path_to_xlsx)


def read_xlsx(path_to_xlsx, worksheet_name, row_range):
    """ Pandas employs openpyxl to read data from .xlsx/.xls
    @param path_to_xlsx: Path to .xlsx;
    @param worksheet_name: Worksheet name;
    @param row_range: Which rows are required?
    @return: Desired rows.
    """
    pddf = pd.read_excel(path_to_xlsx,
                         sheet_name=worksheet_name,
                         row_range=[],
                         header=None
                        )
    return(pddf.loc[row_range].values)


def obtain_parsable_cert_list(tlsimpl):
    """ To obtain parsable certificate file list; called by analyzability_dis_hunter()
    @param tlsimpl: Name of one TLS implementation;
    @return: Parsable certificates.
    """
    file_list = os.listdir(dir_cp_tmp+tlsimpl)
    file_list.remove("cannot_parsed_by_"+tlsimpl.lower()+".txt")
    file_list.sort()
    file_list_cannot_parsed_by_tlsimpl = []
    with open(dir_cp_tmp+tlsimpl+cannot_parsed_by_tlsimpl+tlsimpl.lower()+".txt") as fhr:
        print(dir_cp_tmp+tlsimpl+cannot_parsed_by_tlsimpl)
        for f in fhr.readlines():
            file_list_cannot_parsed_by_tlsimpl.append(os.path.splitext(f)[0].replace('\n', '')+'.'+tlsimpl.lower())
    parsable_file_list = list(set(file_list)-set(file_list_cannot_parsed_by_tlsimpl))
    parsable_file_list.sort()
    return(parsable_file_list)


def obtain_parsability_info(path_to, cannot_parse_file):
    """ To obtain info on non-parsability; called by analyzability_dis_hunter().
    @param path_to: Path to the cannot_parse_file;
    @param cannot_parse_file: A file recording files that cannot be parsed;
    @return: Parsability information. 
    """
    cannot_list = [] 
    cannot_info = [] 
    with open(path_to+cannot_parse_file) as fhr:
        cannot_list = [cert.strip() for cert in fhr.readlines()]
    tmp_list = os.listdir(path_to)
    tmp_list.remove(cannot_parse_file)
    tmp_list.sort() 
    all_list = [os.path.splitext(t)[0] for t in tmp_list]
    
    for c in cannot_list:
        for a in tmp_list:
            if os.path.splitext(a)[0] == c:
                with open(path_to+a) as fhr:
                    cannot_info.append(fhr.read())
    return (all_list, cannot_list, cannot_info)


def analyzability_dis_hunter(path_to, tls_list):
    """ To hunt discrepancies in analyzability.
    @param path_to: Path to TLS folders;
    @param tls_list: A list of TLSs;
    @return: Analyzability discrepancies. 
    """
    global global_cert_list 
    cert_list = []
    can_or_not = []
    error_list = []
    for tls in tls_list:
        dest_path = path_to + tls + '/'
        cannot_parse_file = cannot_parsed_by_tlsimpl + tls.lower() + ".txt"
        all_list, cannot_list, cannot_info = obtain_parsability_info(dest_path, cannot_parse_file)
        tmp_list = ['' for a in all_list]
        for i, c in enumerate(cannot_list):
            id = all_list.index(c)
            tmp_list[id] = cannot_info[i]
        yes_or_no_list = ['y' if t=='' else 'n' for t in tmp_list]
        cert_list.append(all_list)
        can_or_not.append(yes_or_no_list)
        error_list.append(tmp_list)
    for i in range(len(cert_list)-1):
        if cert_list[i] != cert_list[i+1]:
            print("Error: certs list are not comparable!")
            exit()

    global_cert_list = [t for t in cert_list[0]] 
    ncert = len(cert_list[0]) 
    ntls = len(tls_list) 
    discrepancy_list = []
    discrepancy_set = set([])
    unique_discrepancy_list = [] 
    separated_col = ['' for c in range(ncert)]  
    for i in range(ncert):
        encode_str = ''
        for j in range(ntls):
            encode_str += can_or_not[j][i]
        if not encode_str in ['y'*ntls, 'n'*ntls]:
            discrepancy_list.append("discrepancy")
            if not encode_str in discrepancy_set:
                discrepancy_set.add(encode_str)
                unique_discrepancy_list.append("unique")
            else:
                unique_discrepancy_list.append('')
        else:
            discrepancy_list.append('')
            unique_discrepancy_list.append('')
    
    written_dict = {"cert name": cert_list[0],
                    "discrepancy": discrepancy_list,
                    "unique discrepancy": unique_discrepancy_list
                   }
    for i, tls in enumerate(tls_list):
        written_dict[tls] = can_or_not[i]
    for i, tls in enumerate(tls_list):
        written_dict[tls+" error info"] = error_list[i]
    
    worksheet_name = "parsability"
    write_xlsx(file_result, worksheet_name, written_dict, True)

    print("parsability statistics".center(60, '*'))
    print("parsability discrepancy: ", discrepancy_list.count("discrepancy"))
    print("parsability unique discrepancy: ", unique_discrepancy_list.count("unique"))
    return discrepancy_list.count("discrepancy"), unique_discrepancy_list.count("unique")


def get_cert_struc(p):
    """ To obtain structure of a uniform cert; called by get_certs_struc().
    @param p: A parsed cert;
    @return: Certificate structure.
    """
    tp = type(p)
    tmp_dict = {}
    if tp != dict and tp != list:
        return(None)
    elif tp == dict:
        for k, v in p.items():
            struct_v = get_cert_struc(v)
            tmp_dict[k] = struct_v
        return(tmp_dict)
    elif tp == list:
        tmp_list = []
        for v in p:
            struct_v = get_cert_struc(v)
            tmp_list.append(struct_v)
        return(tmp_list)


def get_certs_struc(path_to, tls):
    """ To obtain structure of uniform certs; called by structure_dis_hunter().
    @param path_to: Path to certs;
    @param tls: TLS;
    @return: Certs' structures.
    """
    struc_list = []
    for g in global_cert_list:
        cert = g + '.' + tls.lower()
        path_to_cert = path_to + os.sep + cert
        if os.path.exists(path_to_cert):
            with open(path_to_cert) as fhr:
                cert_value = fhr.read()
            cert_dict = json.loads(cert_value)
            struc_list.append(get_cert_struc(cert_dict))
        else:
            struc_list.append(None)
    return(struc_list)


def dict2jsonstr(d):
    """ To convert dict to JSON str.
    @param d: Dict;
    @return: JSON str.
    """
    def set_default(obj):
        if isinstance(obj, set):
            return(list(obj))
        raise TypeError
    return(json.dumps(d, default=set_default))


def jsonstr2dict(js):
    """ To convert JSON str to dict;
    @param js: JSON str;
    @return: Dict.
    """
    d = json.loads(js)
    return(json.dumps(d, indent=2))


def list2set(list1):
    """ To convert list to set.
    @param list1: List;
    @return: Set.
    """
    set1 = set(list1)
    set2 = set()
    dict1 = {}
    for l in set1:
        dict1[l] = list1.count(l)
        set2.add(l.lower()+'*'*(dict1[l]-1))
    return(set2)


def list2dict_recursive(list1):
    """ To convert list to set (the list may be embedded and the conversion is recursive).
    @param list1: A (possibly embedded) list;
    @return: A dict indicating which elements and their numbers are in the input list.
    """
    dict1 = dict()
    for e in list1:
        if isinstance(e, list):
            dict2 = list2dict_recursive(e)
            for k2, v2 in dict2.items():
                if k2 in dict1.keys():
                    v1 = dict1[k2]
                    dict1[k2] = v1 + v2 
                else:                 
                    dict1[k2] = v2
        elif e.lower() in dict1.keys(): 
            v1 = dict1[e.lower()]
            dict1[e.lower()] = v1 + 1
        elif e != "NA":
            dict1[e.lower()] = 1
        else:                           
            print("e: ", e)
    return(dict1)


def get_structure_dis(struc_lists):
    """ To get structure discrepancies.
    @param struc_lists: Structure Lists e.g., [struc_list_tls1, struc_list_tls2, struc_list_tls3];
    @return: Structure discrepancies. 
    """
    struc_intersection = [] 
    struc_difference_set = []
    tls_num = len(struc_lists)
    dis_list = []
    dis_set = set()
    
    level1_struc_set_list = [] 
    level1_struc_set_list_sift = [] 
    for i in range(tls_num):
        if struc_lists[i]:  
            level1_struc_set_i = set(struc_lists[i].keys()) 
            level1_struc_set_list.append(level1_struc_set_i) 
            level1_struc_set_list_sift.append(level1_struc_set_i) 
        else:
            level1_struc_set_list.append("NA") 

    if len(level1_struc_set_list_sift) in [0, 1]:  
        return(dis_list, struc_intersection, struc_difference_set)

    level1_struc_intersection = set()
    for i in range(len(level1_struc_set_list_sift)):
        if i == 0:  
            level1_struc_intersection = level1_struc_set_list_sift[i] 
        else:
            level1_struc_intersection = level1_struc_intersection & level1_struc_set_list_sift[i]
    struc_intersection.append(level1_struc_intersection) 

    difference_set_list = [] 
    for i in range(tls_num):
        if not level1_struc_set_list[i] in ["NA"]: 
            difference_set = level1_struc_set_list[i] - level1_struc_intersection
            difference_set_list.append(difference_set) 
        else:
            difference_set_list.append("NA") 
    struc_difference_set.append(difference_set_list)
    
    dis_flag = False
    for i in range(tls_num):
        if not difference_set_list[i] in [set(), "NA"]:
            dis_flag = True
    if dis_flag:
        dis_list.append(difference_set_list)
    
    level2_struc_set_list = []
    level2_struc_set_list_sift = []
    for i in range(tls_num):
        if struc_lists[i]: 
            for k in level1_struc_intersection:
                if type(struc_lists[i][k]) == dict:
                    level2_keys = struc_lists[i][k].keys() 
                    level2_struc_set_i = set(level2_keys) 
                    level2_struc_set_list.append(level2_struc_set_i)
                    level2_struc_set_list_sift.append(level2_struc_set_i)
                else:
                    pass
        else:
            level2_struc_set_list.append("NA")

    if len(level2_struc_set_list_sift) in [0, 1]:
        return(dis_list, struc_intersection, struc_difference_set) 

    level2_struc_intersection = set()
    for i in range(len(level2_struc_set_list_sift)):
        if i == 0:
            level2_struc_intersection = level2_struc_set_list_sift[i]
        else:
            level2_struc_intersection = level2_struc_intersection & level2_struc_set_list_sift[i]
    struc_intersection.append(level2_struc_intersection)

    difference_set_list = [] 
    for i in range(tls_num):
        if not level2_struc_set_list[i] in ["NA"]: 
            difference_set = level2_struc_set_list[i] - level2_struc_intersection
            difference_set_list.append(difference_set)
        else:
            difference_set_list.append("NA")
    struc_difference_set.append(difference_set_list)
    
    dis_flag = False
    for i in range(tls_num):
        if not difference_set_list[i] in [set(), "NA"]:
            dis_flag = True
    if dis_flag:
        dis_list.append(difference_set_list)

    level3_struc_set_list = []
    level3_struc_set_list_sift = []
    for i in range(tls_num):
        level3_struc_pki_set = set() 
        level3_struc_ext_set_list = []
        level3_struc_ext_set = set()
        if struc_lists[i]: 
            if "subject public key info" in level2_struc_intersection:
                for k in struc_lists[i]["tbsCertificate"]["subject public key info"].keys():
                    level3_struc_pki_set.add(k.lower())
            else:
                level3_struc_pki_set = set()

            if "X509v3 extensions" in level2_struc_intersection:
                level3_struc_ext_set_list = struc_lists[i]["tbsCertificate"]["X509v3 extensions"]
            else:
                level3_struc_ext_set_list = []

            tmp_list = [] 
            for d in level3_struc_ext_set_list:
                for ext in d.keys():
                    tmp_list.append(ext)
            
            level3_struc_ext_set = list2set(tmp_list)
            
            level3_struc_set_i = level3_struc_pki_set | level3_struc_ext_set 
            level3_struc_set_list.append(level3_struc_set_i)
            level3_struc_set_list_sift.append(level3_struc_set_i)
        else:
            level3_struc_set_list.append("NA")

    if len(level3_struc_set_list_sift) in [0, 1]:
        return(dis_list, struc_intersection, struc_difference_set)

    level3_struc_intersection = set()
    for i in range(len(level3_struc_set_list_sift)):
        if i == 0:
            level3_struc_intersection = level3_struc_set_list_sift[i]
        else:
            level3_struc_intersection = level3_struc_intersection & level3_struc_set_list_sift[i]
    struc_intersection.append(level3_struc_intersection)

    difference_set_list = []
    for i in range(tls_num):
        if not level3_struc_set_list[i] in ["NA"]: 
            difference_set = level3_struc_set_list[i] - level3_struc_intersection
            difference_set_list.append(difference_set)
        else:
            difference_set_list.append("NA")
    struc_difference_set.append(difference_set_list)
    
    dis_flag = False
    for i in range(tls_num):
        if not difference_set_list[i] in [set(), "NA"]:
            dis_flag = True
    if dis_flag:
        dis_list.append(difference_set_list)

    level4_struc_set_list = []
    level4_struc_set_list_sift = []
    for i in range(tls_num):
        level4_struc_subext_set_list = []
        level4_struc_subext_set = set()
        if struc_lists[i]: 
            if "X509v3 extensions" in level2_struc_intersection: 
                ext_dict_list = struc_lists[i]["tbsCertificate"]["X509v3 extensions"]
                tmp_list = [] 
                for ext_no, ext_dict in enumerate(ext_dict_list):   
                    if type(ext_dict) == dict: 
                        for k, v in ext_dict.items(): 
                            if type(v) == list:
                                for subext in v:
                                    if type(subext) == dict: 
                                        tmp_list.extend(subext.keys())
                                    else:
                                        pass
                            elif type(v) == dict:
                                tmp_list.extend(v.keys())
                            else:
                                pass
                    else:
                        print("Non-dict ext found!")
                level4_struc_subext_set_list = tmp_list
            else:
                level4_struc_subext_set_list = []   
            
            level4_struc_subext_set = list2set(level4_struc_subext_set_list)
            
            level4_struc_set_i = level4_struc_subext_set
            level4_struc_set_list.append(level4_struc_set_i)
            level4_struc_set_list_sift.append(level4_struc_set_i)
        else:
            level4_struc_set_list.append("NA")
    
    if len(level4_struc_set_list_sift) in [0, 1]:
        return(dis_list, struc_intersection, struc_difference_set)

    level4_struc_intersection = set()
    for i in range(len(level4_struc_set_list_sift)):
        if i == 0:
            level4_struc_intersection = level4_struc_set_list_sift[i]
        else:
            level4_struc_intersection = level4_struc_intersection & level4_struc_set_list_sift[i]
    struc_intersection.append(level4_struc_intersection)

    difference_set_list = []
    for i in range(tls_num):
        if not level4_struc_set_list[i] in ["NA"]: 
            difference_set = level4_struc_set_list[i] - level4_struc_intersection
            difference_set_list.append(difference_set)
        else:
            difference_set_list.append("NA")
    struc_difference_set.append(difference_set_list)
    
    dis_flag = False
    for i in range(tls_num):
        if not difference_set_list[i] in [set(), "NA"]:
            dis_flag = True
    if dis_flag:
        dis_list.append(difference_set_list)

    level5_struc_set_list = []
    level5_struc_set_list_sift = []
    for i in range(tls_num):
        level5_struc_subsubext_set_list = []
        level5_struc_subsubext_set = set()
        if struc_lists[i]:
            if "X509v3 extensions" in level2_struc_intersection:
                ext_dict_list = struc_lists[i]["tbsCertificate"]["X509v3 extensions"] 
                tmp_list = [] 
                for ext_no, ext_dict in enumerate(ext_dict_list):
                    if type(ext_dict) == dict:
                        for k, v in ext_dict.items(): 
                            if type(v) == list:
                                for subext in v:
                                    if type(subext) == dict:
                                        for kk, vv in subext.items():
                                            if type(vv) == list:
                                                for subsubext in vv:
                                                    if type(subsubext) == dict:
                                                        tmp_list.extend(subsubext.keys())
                                                    else:
                                                        pass
                                            elif isinstance(vv, dict):
                                                tmp_list.extend(vv.keys())
                                            else:
                                                pass
                                    else:
                                        pass
                            else:
                                pass
                    else:
                        print("Non-dict ext found!")
                level5_struc_subext_set_list = tmp_list
            else:
                level5_struc_subext_set_list = []
            
            level5_struc_subext_set = list2set(level5_struc_subext_set_list)

            level5_struc_set_i = level5_struc_subext_set
            level5_struc_set_list.append(level5_struc_set_i)
            level5_struc_set_list_sift.append(level5_struc_set_i)
        else:
            level5_struc_set_list.append("NA")

    if len(level5_struc_set_list_sift) in [0, 1]:
        return(dis_list, struc_intersection, struc_difference_set)

    level5_struc_intersection = set()
    for i in range(len(level5_struc_set_list_sift)):
        if i == 0:
            level5_struc_intersection = level5_struc_set_list_sift[i]
        else:
            level5_struc_intersection = level5_struc_intersection & level5_struc_set_list_sift[i]
    struc_intersection.append(level5_struc_intersection)

    difference_set_list = [] 
    for i in range(tls_num):
        if not level5_struc_set_list[i] in ["NA"]: 
            difference_set = level5_struc_set_list[i] - level5_struc_intersection
            difference_set_list.append(difference_set)
        else:
            difference_set_list.append("NA")
    struc_difference_set.append(difference_set_list)
    
    dis_flag = False
    for i in range(tls_num):
        if not difference_set_list[i] in [set(), "NA"]:
            dis_flag = True
    if dis_flag:
        dis_list.append(difference_set_list)

    level6_struc_set_list = []
    level6_struc_set_list_sift = []
    for i in range(tls_num):
        level6_struc_subsubext_set_list = []
        level6_struc_subsubext_set = set()
        if struc_lists[i]:
            if "X509v3 extensions" in level2_struc_intersection:
                ext_dict_list = struc_lists[i]["tbsCertificate"]["X509v3 extensions"]
                tmp_list = [] 
                for ext_no, ext_dict in enumerate(ext_dict_list):
                    if type(ext_dict) == dict:
                        for k, v in ext_dict.items():
                            if type(v) == list:
                                for subext in v:
                                    if type(subext) == dict:
                                        for kk, vv in subext.items():
                                            if type(vv) == list:
                                                for subsubext in vv:
                                                    if type(subsubext) == dict:
                                                        for kkk, vvv in subsubext.items():
                                                            if type(vvv) == list:
                                                                for subsubsubext in vvv:
                                                                    if type(subsubsubext) == dict:
                                                                        tmp_list.extend(subsubsubext.keys())
                                                                    elif isinstance(subsubsubext, list):
                                                                        pass
                                                            elif type(vvv) == dict:
                                                                tmp_list.extend(vvv.keys())
                                                            else:
                                                                pass
                                                    else:
                                                        pass
                                            else:
                                                pass
                                    else:
                                        pass
                            else:
                                pass
                    else:
                        pass
                level6_struc_subext_set_list = tmp_list
            else:
                level6_struc_subext_set_list = []
            
            level6_struc_subext_set = list2set(level6_struc_subext_set_list)

            level6_struc_set_i = level6_struc_subext_set
            level6_struc_set_list.append(level6_struc_set_i)
            level6_struc_set_list_sift.append(level6_struc_set_i)
        else:
            level6_struc_set_list.append("NA")

    if len(level6_struc_set_list_sift) in [0, 1]:
        return(dis_list, struc_intersection, struc_difference_set)

    level6_struc_intersection = set()
    for i in range(len(level6_struc_set_list_sift)):
        if i == 0:
            level6_struc_intersection = level6_struc_set_list_sift[i]
        else:
            level6_struc_intersection = level6_struc_intersection & level6_struc_set_list_sift[i]
    struc_intersection.append(level6_struc_intersection)

    difference_set_list = [] 
    for i in range(tls_num):
        if not level6_struc_set_list[i] in ["NA"]: 
            difference_set = level6_struc_set_list[i] - level6_struc_intersection
            difference_set_list.append(difference_set)
        else:
            difference_set_list.append("NA")
    struc_difference_set.append(difference_set_list) 
    
    dis_flag = False
    for i in range(tls_num):
        if not difference_set_list[i] in [set(), "NA"]:
            dis_flag = True
    if dis_flag:
        dis_list.append(difference_set_list) 

    return(dis_list, struc_intersection, struc_difference_set)


def set_default(obj):
    """ code for JSON
    @param obj: Object;
    @return: Default list. 
    """
    if isinstance(obj, set):
        return(list(obj))
    raise TypeError


def structure_dis_hunter(path_to, tls_list):
    """ To hunt discrepancies in structure; note: discrepancies can be found among structures of at least two parsable certs.
    @param path_to: Path to uniform certs;
    @param tls_list: TLS lists;
    @return: Structure discrepancies.
    """
    struc_intersection_list, struc_dis_all_list = [], []
    struc_list = [] 
    for tls in tls_list:
        dest_path = path_to + tls
        struc_list.append(get_certs_struc(dest_path, tls))  

    ncert = len(global_cert_list) 
    separated_col = ['' for c in range(ncert)] 
    struc_unique_dis_list = []
    struc_unique_dis_cert_list = []
    struc_dis_list = [] 
    struc_all_dis_list = []  

    for cert_no, cert_name in enumerate(global_cert_list):
        current_struc_list = [] 
        for tls_no, tls_name in enumerate(tls_list):
            current_struc_list.append(struc_list[tls_no][cert_no])
        struc_dis, struc_intersection, struc_dis_all = get_structure_dis(current_struc_list)

        current_unique_dis = [] 
        for dis_no, dis in enumerate(struc_dis):
            if not (dis in struc_unique_dis_list): 
                struc_unique_dis_list.append(dis) 
                current_unique_dis.append(dis) 
            struc_all_dis_list.append(dis) 
        
        struc_unique_dis_cert_list.append(current_unique_dis)
        struc_dis_list.append(struc_dis)
        n0 = len(struc_unique_dis_list) 
        n1 = len(struc_unique_dis_cert_list) 
        n2 = len(struc_dis_list) 
        n3 = len(struc_all_dis_list) 
        struc_intersection_list.append(struc_intersection) 
        struc_dis_all_list.append(struc_dis_all) 

    written_dict = {}
    written_dict = {"cert name": global_cert_list,
               "struc discrepancy of current cert": struc_dis_list,
               "unique struc discrepancy of current cert": struc_unique_dis_cert_list,
               }
    for i, tls in enumerate(tls_list):
        col_name = "struc parsed by " + tls
        written_dict[col_name] = struc_list[i]
        
    worksheet_name = "structure"
    write_xlsx(file_result, worksheet_name, written_dict)

    print("structure discrepancy statistics".center(60, '*'))
    print("structure discrepancy: ", n3)
    print("structure unique discrepancy: ", n0)
    
    return n3, n0, struc_intersection_list, struc_dis_all_list


def get_certs_value(path_to, tls):
    """ To obtain a list of values of certs for one TLS implementation.
    @param path_to: Path to uniform certs (one TLS implementation);
    @param tls: TLS (one TLS implementation; its named is used instead);
    @return: A list of cert values, which are in the format of dict; the list corresponds to the list of certs.
    """
    value_list = []
    for cert_name_prefix in global_cert_list:
        cert_name = cert_name_prefix + '.' + tls.lower()
        path_to_cert = path_to + os.sep + cert_name
        if os.path.exists(path_to_cert):
            with open(path_to_cert) as fhr:
                cert_value = fhr.read()
            value_dict = json.loads(cert_value)
            value_list.append(value_dict)
        else:
            value_list.append(None) 
    return(value_list)


def determine_list_duplicate(list1):
    """ To determine whether a list contains duplicate elements.
    @param list1: A list;
    @return: Duplicate or not.
    """
    for e in list1:
        if list1.count(e) > 1:
            return("Duplicate")
        else:
            return("noDuplicate")


def get_value_dis(value_dict_tlslist, struc_intersection_certlevellist, struc_dis_tlslist_certlevellist):
    """ To obtain value discrepancies among parsed results for one cert.
    @param value_dict_tlslist: A list of one cert value parsed by multiple TLS implementations, e.g., [cert_value_parsedby_tls1, cert_value_parsedby_tls2, ...] (each element of the list is a dict);
    @param struc_intersection_certlevellist: A list of structure intersection and each element stands for one level of one cert, e.g., [{struc_intersection_level1}, {struc_intersection_level2}, ...];
    @param struc_dis_tlslist_certlevellist: A list of structure discrepancy list at different levels, e.g., [ [{struc_dis_level1_tls1}, {struc_dis_level1_tls2}, {struc_dis_level1_tls3}], [{struc_dis_level2_tls1}, {struc_dis_level2_tls2}, {struc_dis_level2_tls3}],... ];
    @return: Value discrepancies. 
    """
    struc_intersection_value_dis_list = []
    
    struc_difset_value_dis_list = []

    ntls = len(value_dict_tlslist) 
    
    nlevel = len(struc_intersection_certlevellist)
    
    cur_and_lower_level_value_tlslist = []
    
    for cur_level in range(nlevel):
        cur_level_value_tlslist = []
        
        lower_level_value_tlslist = []
        
        if cur_level == 0:
            cur_and_lower_level_value_tlslist = value_dict_tlslist
        else:
            cur_and_lower_level_value_tlslist = cur_and_lower_level_value_tlslist
        
        for cur_tls in range(ntls):
            cur_and_lower_level_value_xtls = cur_and_lower_level_value_tlslist[cur_tls]
            
            cur_level_value_xtls = []   
            lower_level_value_xtls = [] 
            
            if isinstance(cur_and_lower_level_value_xtls, dict):
                for k, v in cur_and_lower_level_value_xtls.items(): 
                    if isinstance(v, dict):
                        cur_level_value_xtls.append({k:"value-contain-dict-type"})  
                        for kk, vv in v.items():
                            lower_level_value_xtls.append({kk:vv})  
                    elif isinstance(v, list):
                        if v:   
                            flag_list_has_dict = False  
                            non_dict_element_list = []  
                            for e in v: 
                                if isinstance(e, dict):
                                    flag_list_has_dict = True
                                    for kk, vv in v.items():
                                        lower_level_value_xtls.append({kk:vv})  
                                else:   
                                    non_dict_element_list.append(e) 
                            if flag_list_has_dict:
                                non_dict_element_list.append("value-contain-dict-type") 
                                cur_level_value_xtls.append({k:non_dict_element_list})  
                            else:   
                                cur_level_value_xtls.append({k:v})  
                        else:   
                            cur_level_value_xtls.append({k:list()}) 
                    else:   
                        cur_level_value_xtls.append({k:v})  
            elif isinstance(cur_and_lower_level_value_xtls, list):  
                if cur_and_lower_level_value_xtls:  
                    for e in cur_and_lower_level_value_xtls:    
                        if isinstance(e, dict): 
                            for k, v in e.items():  
                                if isinstance(v, dict): 
                                    cur_level_value_xtls.append({k:"value-contain-dict-type"})
                                    for kk, vv in v.items():
                                        lower_level_value_xtls.append({kk:vv})  
                                elif isinstance(v, list):      
                                    if v:   
                                        flag_list_has_dict = False
                                        non_dict_element_list = []
                                        for ee in v:
                                            if isinstance(ee, dict):
                                                flag_list_has_dict = True
                                                for kk, vv in ee.items():
                                                    lower_level_value_xtls.append({kk:vv})
                                            else:
                                                non_dict_element_list.append(ee)
                                        if flag_list_has_dict:
                                            non_dict_element_list.append("value-contain-dict-type")
                                            cur_level_value_xtls.append({k:non_dict_element_list})
                                        else:   
                                            cur_level_value_xtls.append({k:v})
                                    else:
                                        cur_level_value_xtls.append({k:list()}) 
                                else:   
                                    cur_level_value_xtls.append({k:v})
                        elif isinstance(e, list):   
                            if e:
                                pass
                            else:
                                pass
                        else:   
                            pass
                else:   
                    pass 
            else:   
                pass
            cur_level_value_tlslist.append(cur_level_value_xtls)
            lower_level_value_tlslist.append(lower_level_value_xtls)
        
        cur_level_struc_intersection = struc_intersection_certlevellist[cur_level]
        
        cur_level_struc_dis_tlslist = struc_dis_tlslist_certlevellist[cur_level]
        
        cur_level_struc_intersection_value_dis_tlslist = []
        
        if cur_level_struc_intersection:
            for cur_struc in cur_level_struc_intersection:
                cur_struc_value_tlslist = [] 
                for cur_tls in range(ntls):
                    cur_level_value_curtls = cur_level_value_tlslist[cur_tls]
                    if isinstance(cur_level_value_curtls, dict):
                        for k in cur_level_value_curtls.keys():
                            if cur_struc.lower() == k.lower():  
                                cur_struc_value_tlslist.append(cur_level_value_curtls[k])
                                break
                        else:
                            cur_struc_value_tlslist.append("NA") 
                    elif isinstance(cur_level_value_curtls, list):
                        cur_struc_valuelist = [] 
                        for e in cur_level_value_curtls:
                            if isinstance(e, dict):
                                for k in e.keys():
                                    if cur_struc.lower() == k.lower():
                                        if cur_struc == "pathlen":
                                            cur_struc_valuelist.append(int(e[k]))
                                        elif isinstance(e[k], str):
                                            cur_struc_valuelist.append(e[k].lower())
                                        elif isinstance(e[k], list):
                                            tmp_list = []
                                            for ee in e[k]:
                                                if isinstance(ee, str):
                                                    tmp_list.append(ee.lower())
                                                else:
                                                    tmp_list.append(ee)
                                            cur_struc_valuelist.append(tmp_list)
                                        else:
                                            cur_struc_valuelist.append(e[k])
                            elif isinstance(e, list):
                                pass 
                            else:
                                cur_struc_valuelist.append(e)
                        nelement = len(cur_struc_valuelist) 
                        if nelement == 0: 
                            cur_struc_value_tlslist.append("NA")
                        elif nelement == 1:
                            cur_struc_value_tlslist.append(cur_struc_valuelist[0])
                        else:
                            tmp_list = []
                            for e in cur_struc_valuelist:
                                if isinstance(e, str):
                                    tmp_list.append(e.lower())
                                else:
                                    tmp_list.append(e)
                            cur_struc_value_tlslist.append(tmp_list)
                    else:  
                        cur_struc_value_tlslist.append(cur_level_value_curtls)
            
                cur_struc_value_dis = []
                flag_cur_struc_value_dis = False
                value_x = cur_struc_value_tlslist[0]
                value_x_type = type(value_x)
                cur_struc_value_dis.append(value_x)
                for next_tls in range(1, ntls):
                    value_y = cur_struc_value_tlslist[next_tls]
                    value_y_type = type(value_y)
                    if not flag_cur_struc_value_dis:
                        if value_x_type == list and value_y_type == list:
                            for e1 in value_x:
                                if not e1 in value_y:
                                    flag_cur_struc_value_dis = True
                                else:
                                    pass
                            for e2 in value_y:
                                if not e2 in value_x:
                                    flag_cur_struc_value_dis = True
                                else:
                                    pass
                        elif value_x_type == list and value_y_type != list:
                            if len(value_x) == 1:
                                if value_x[0] == value_y:
                                    pass
                                else:
                                    flag_cur_struc_value_dis = True
                            else:
                                flag_cur_struc_value_dis = True
                        elif value_x_type != list and value_y_type == list:
                            if len(value_y) == 1:
                                if value_y[0] == value_x:
                                    pass
                                else:
                                    flag_cur_struc_value_dis = True
                            else:
                                flag_cur_struc_value_dis = True
                        else:
                            if value_x == value_y:
                                pass
                            else:
                                flag_cur_struc_value_dis = True
                    else: 
                        pass
                    cur_struc_value_dis.append(value_y)
                    
                if not flag_cur_struc_value_dis:
                    cur_struc_value_dis = []
                if cur_struc_value_dis:
                    cur_level_struc_intersection_value_dis_tlslist.append(cur_struc_value_dis)
        else: 
            pass
        
        struc_intersection_value_dis_list.append(cur_level_struc_intersection_value_dis_tlslist)
        
        if cur_level_struc_dis_tlslist:
            cur_level_struc_dis_count_dict_tlslist = []
            for cur_level_struc_dis in cur_level_struc_dis_tlslist:
                if cur_level_struc_dis: 
                    tmp_dict = dict()
                    for e in cur_level_struc_dis:
                        e_count = e.count('*') + 1
                        e = re.sub('\*', '', e)
                        tmp_dict.update({e:e_count})
                    cur_level_struc_dis_count_dict_tlslist.append(tmp_dict)
                else:   
                    cur_level_struc_dis_count_dict_tlslist.append(dict())  
            
            cur_level_struc_dis_value_dis = []
            flag_cur_level_struc_dis_value_dis = False
            
            cur_level_struc_dis_occur_dict = dict()
            for cur_tls in range(ntls):
                cur_level_struc_dis_count_dict = cur_level_struc_dis_count_dict_tlslist[cur_tls]
                for key in cur_level_struc_dis_count_dict:
                    if key in cur_level_struc_dis_occur_dict.keys():
                        value = cur_level_struc_dis_occur_dict[key] 
                        value.add(cur_tls) 
                        cur_level_struc_dis_occur_dict.update({key:value})
                    else:  
                        cur_level_struc_dis_occur_dict.update({key:set([cur_tls])})

            cur_level_struc_difset_value_dis_list = list()
            if cur_level_struc_dis_occur_dict:
                for cur_struc, cur_struc_occur_set in cur_level_struc_dis_occur_dict.items():
                    if len(cur_struc_occur_set) > 1: 
                        cur_struc_value_tlslist = [] 
                        for cur_tls in range(ntls):
                            if cur_tls in cur_struc_occur_set:
                                cur_level_value_curtls = cur_level_value_tlslist[cur_tls]
                                tmp_list = [] 
                                for e in cur_level_value_curtls:
                                    if isinstance(e, dict):
                                        for ek in e.keys():
                                            if cur_struc.lower() == ek.lower():
                                                tmp_list.append(e[ek])
                                    else: 
                                        pass
                                if len(tmp_list) == 1:
                                    cur_struc_value_tlslist.append(tmp_list[0])
                                elif len(tmp_list) > 1:
                                    cur_struc_value_tlslist.append(tmp_list)
                                else:
                                    pass
                            else: 
                                cur_struc_value_tlslist.append("NA")
                        
                        cur_struc_value_dict_tlslist = []
                        for cur_struc_value in cur_struc_value_tlslist:
                            cur_struc_value_dict_curtls = dict()
                            if isinstance(cur_struc_value, list):
                                cur_struc_value_dict_curtls = list2dict_recursive(cur_struc_value)
                            elif isinstance(cur_struc_value, str) and cur_struc_value != "NA": 
                                cur_struc_value_dict_curtls = {cur_struc_value.lower():1}
                            elif cur_struc_value != "NA":
                                cur_struc_value_dict_curtls = {cur_struc_value:1}
                            cur_struc_value_dict_tlslist.append(cur_struc_value_dict_curtls)
                        
                        cur_struc_value_dis_tlslist = []
                        for cur_tls in range(ntls):
                            if cur_tls in cur_struc_occur_set and cur_tls < len(cur_struc_value_dict_tlslist):
                                cur_struc_value_dis_curtls = dict() 
                                cur_struc_value_dict_curtls = cur_struc_value_dict_tlslist[cur_tls]
                                for cur_struc_cur_value_curtls, cur_struc_cur_value_curtls_count in cur_struc_value_dict_curtls.items():
                                    for other_tls in range(ntls):
                                        if cur_tls != other_tls: 
                                            if other_tls in cur_struc_occur_set and other_tls < len(cur_struc_value_dict_tlslist):
                                                cur_struc_value_dict_othertls = cur_struc_value_dict_tlslist[other_tls]
                                                if cur_struc_cur_value_curtls in cur_struc_value_dict_othertls.keys():
                                                    cur_struc_cur_value_othertls = cur_struc_cur_value_curtls
                                                    cur_struc_cur_value_othertls_count = cur_struc_value_dict_othertls[cur_struc_cur_value_othertls]
                                                    if cur_struc_cur_value_curtls_count != cur_struc_cur_value_othertls_count:
                                                        pass
                                                    else:
                                                        pass
                                                else: 
                                                    cur_struc_value_dis_curtls[cur_struc_cur_value_curtls] = cur_struc_cur_value_curtls_count
                                            else: 
                                                pass
                                        else: 
                                            pass
                                cur_struc_value_dis_tlslist.append(cur_struc_value_dis_curtls)
                            else: 
                                cur_struc_value_dis_tlslist.append("NA")
                        
                        flag_cur_struc_value_dis_tlslist = False
                        for e in cur_struc_value_dis_tlslist:
                            if e in [dict(), "NA"]:
                                pass
                            else:
                                flag_cur_struc_value_dis_tlslist = True
                        if flag_cur_struc_value_dis_tlslist:
                            if cur_struc_value_dis_tlslist not in cur_level_struc_difset_value_dis_list:
                                cur_level_struc_difset_value_dis_list.append(cur_struc_value_dis_tlslist)
                    else: 
                        pass
                struc_difset_value_dis_list.append(cur_level_struc_difset_value_dis_list)
            else: 
                struc_difset_value_dis_list.append("NA")
            
        else: 
            struc_difset_value_dis_list.append(list())     
        
        cur_and_lower_level_value_tlslist = lower_level_value_tlslist
        
    return struc_intersection_value_dis_list, struc_difset_value_dis_list


def value_dis_hunter(path_to, tls_list, struc_intersection_list, struc_dis_all_list):
    """ To hunt discrepancies in value.
    @param path_to: Path to uniform certs;
    @param tls_list: TLS list;
    @return: Value discrepancies.
    """

    value_list = []
    for tls in tls_list:
        dest_path = path_to + tls
        value_list.append(get_certs_value(dest_path, tls))
    
    separated_col = ['' for c in range(len(global_cert_list))]
    value_notunique_dis_list_allcertsdis = [] 
    value_notunique_dis_list_eachcert = [] 
    value_unique_dis_list_allcertsdis = [] 
    value_unique_dis_list_eachcert = [] 
    struc_intersection_value_dis_list_allcertslist = []
    struc_difset_value_dis_list_allcertslist = []
    
    for cert_no, cert_name in enumerate(global_cert_list):
        current_value_list = []
        for tls_no, tls_name in enumerate(tls_list):
            current_value_list.append(value_list[tls_no][cert_no])
        struc_intersection_value_dis_list, struc_difset_value_dis_list = get_value_dis(current_value_list,
                                           struc_intersection_list[cert_no],
                                           struc_dis_all_list[cert_no])
        
        struc_intersection_value_dis_list_allcertslist.append(struc_intersection_value_dis_list)
        struc_difset_value_dis_list_allcertslist.append(struc_difset_value_dis_list)
        
        value_notunique_dis_list_curcert = []
        value_unique_dis_list_curcert = []
        
        for cur_level_value_dis_list in struc_intersection_value_dis_list:
            if cur_level_value_dis_list not in ["NA", list()]:
                for cur_level_value_dis in cur_level_value_dis_list:
                    if cur_level_value_dis not in ["NA", list()]:
                        value_notunique_dis_list_allcertsdis.append(cur_level_value_dis)
                        if cur_level_value_dis not in value_unique_dis_list_allcertsdis:
                            value_unique_dis_list_allcertsdis.append(cur_level_value_dis)
                        value_notunique_dis_list_curcert.append(cur_level_value_dis)
                        if cur_level_value_dis not in value_unique_dis_list_curcert:
                            value_unique_dis_list_curcert.append(cur_level_value_dis)
                    
        for cur_level_value_dis_list in struc_difset_value_dis_list:
            if cur_level_value_dis_list not in ["NA", list()]:
                for cur_level_value_dis in cur_level_value_dis_list:
                    if cur_level_value_dis not in ["NA", list()]:
                        value_notunique_dis_list_allcertsdis.append(cur_level_value_dis)
                        if cur_level_value_dis not in value_unique_dis_list_allcertsdis:
                            value_unique_dis_list_allcertsdis.append(cur_level_value_dis)
                        value_notunique_dis_list_curcert.append(cur_level_value_dis)
                        if cur_level_value_dis not in value_unique_dis_list_curcert:
                            value_unique_dis_list_curcert.append(cur_level_value_dis)
        
        value_notunique_dis_list_eachcert.append(value_notunique_dis_list_curcert)
        value_unique_dis_list_eachcert.append(value_unique_dis_list_curcert)    
        
        num_value_notunique_dis_list_eachcert = len(value_notunique_dis_list_curcert)
        num_value_unique_dis_list_eachcert = len(value_unique_dis_list_curcert)
    
    num_value_notunique_dis_list_allcertsdis = len(value_notunique_dis_list_allcertsdis)
    num_value_unique_dis_list_allcertsdis = len(value_unique_dis_list_allcertsdis) 
    
    written_dict = {}
    written_dict = {"cert name": global_cert_list,
               "value discrepancy of current cert": value_notunique_dis_list_eachcert,
                "unique value discrepancy of current cert": value_unique_dis_list_eachcert
               }
    for i, tls in enumerate(tls_list):
        col_name = "value parsed by " + tls
        written_dict[col_name] = value_list[i]
    
    worksheet_name = "value"
    write_xlsx(file_result, worksheet_name, written_dict)

    print("value discrepancy statistics".center(60, '*'))
    print("value discrepancy: ", num_value_notunique_dis_list_allcertsdis)
    print("value unique discrepancy: ", num_value_unique_dis_list_allcertsdis)
    
    return num_value_notunique_dis_list_allcertsdis, num_value_unique_dis_list_allcertsdis
    

def test_open_cert(path_to, cert_name):
    """ To test open certs.
    @param path_to: Path to certs;
    @param cert_name: Cert name;
    @return: Test information.
    """
    with open(path_to+cert_name) as fhr:
        print(fhr.read())


def SBDT_discrepancy_hunter_main():
    """ To hunt discrepancies of analyzability, structure, and value.
    @param None: None;
    @return: Discrepancies of analyzability, structure, and value.
    """
    begin_time = time.time()
    
    num_pd, num_dpd = analyzability_dis_hunter(dir_cp_tmp, tlsimpl)
    
    num_struc_dis, num_unique_struc_dis, struc_intersection_list, struc_dis_all_list = structure_dis_hunter(dir_cp_uniform, tlsimpl)
    
    num_value_dis, num_unique_value_dis = value_dis_hunter(dir_cp_uniform, tlsimpl, struc_intersection_list, struc_dis_all_list)
    
    end_time = time.time()
    
    print("SBDT_4_discrepancy_hunter succeeds!".center(60, '*'))
    print(("Time elapsed: "+str(end_time-begin_time)+" seconds.").center(60, '*'))
    return num_pd, num_dpd, num_struc_dis, num_unique_struc_dis, num_value_dis, num_unique_value_dis


if __name__ == "__main__":
    SBDT_discrepancy_hunter_main()
    